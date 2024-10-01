import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
import numpy as np
import os
from copy import deepcopy
os.chdir(os.path.dirname(__file__))
import gc
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from datetime import datetime


st.set_page_config(page_title="Prasa w mediach społecznościowych", page_icon=":book:")
st.markdown("<h1 style='margin-top: -80px; text-align: center;'>Prasa w mediach społecznościowych</h1>", unsafe_allow_html=True)

my_colors = {
    'Facebook': '#0070C0',
    'YouTube': '#981923',
    'X': '#193441',
    'LinkedIn': '#006683',
    'TikTok': '#00F0F0',
    'Pinterest': '#E0404B',
    'Instagram': "#5B0F15",
    'Suma': '#0F1F27'
}

table_css_style = """
    <style>
        table.sticky-header thead tr {
            position: sticky;
            top: 0;
            background-color: #f0f2f6;
            border: 0.2em solid ##77AADB;
        }

        table.sticky-header {
            font-size: 0.79em; /* Adjust the font size as needed */
            max-width: 100%; /* Adjust the width as needed */
            overflow-x: auto; /* Add horizontal scroll if needed */
        }

        /* Add a new style for the first column */
        table.sticky-header td.col0 {
            text-align: center;
        }

        td {
        white-space: nowrap; /* This prevents text from breaking into multiple lines */
        }
    </style>
"""


@st.cache_data(ttl=3600)
def load_data(filename, indexcol=False):
    if not indexcol:
        df = pd.read_excel(filename)
    else:
        df = pd.read_excel(filename, index_col=0)
    return df


def format_number_with_spaces(number_str):
    # formatowanie liczb do formatu 1 111 111
    reversed_str = str(int(number_str))[::-1]
    groups = [reversed_str[i:i + 3] for i in range(0, len(reversed_str), 3)]
    return ' '.join(groups)[::-1]
    

######## Ładowanie danych ########
# dane dot. followersów
df = load_data('./df_followers.xlsx', indexcol=True)

# dane dot. periodyczności
mapa = load_data('./mapa_typy_pism.xlsx')
df = df.merge(mapa, on='Tytuł', how='left')
df = df[df['Typ']!='NIEUWZGLĘDNIONE']
df.set_index(df.columns[0], inplace=True)
donutdf =deepcopy(df.drop(['Typ'], axis=1)) # ramka danych do donut charta

# dane dot. hyperlinków
hyper_df = load_data('./mapa_adresy_pbc.xlsx')
hyperlink_dict = {title: address for title, address in zip(hyper_df['Tytuł'], hyper_df['AdresPBC'])}

def add_hyperlink(value, hyperlink_dict=hyperlink_dict):
    if value in hyperlink_dict:
        return f'<a href="{hyperlink_dict[value]}" style="text-decoration:none; color:black;">{value}</a>'
    return value


######## Wykresy ########

@st.cache_data(hash_funcs={matplotlib.figure.Figure: lambda _: None}, ttl=3600)
def barplot(df, column):
    fig, ax = plt.subplots(figsize=(8, 6))
    plt.barh(df.index, df, color=my_colors[column], height=0.5)
    plt.gca().spines[:].set_visible(False)
    plt.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False)
    plt.tick_params(axis='y', which='both', length=0, labelleft=False) # labelleft=False żeby wykresy zaczynały się w tym samym miejscu
    plt.gca().invert_yaxis()

    plt.title(f'Top 10 {column}', loc='left', fontdict={'fontsize': 14, 'fontweight': 'bold', 'fontname': 'Lato'})

    margin = max(df)*0.02
    for index, value in enumerate(list(df)):
        if value>0:
            plt.text(value+margin, index+.1, format_number_with_spaces(value))
        
    for index, pismo in enumerate(list(df.index)):
        plt.text(0, index-0.48, pismo, ha='left', va='center', fontdict={'fontsize': 10.8, 'fontname': 'Lato'})
    return fig


@st.cache_data(hash_funcs={matplotlib.figure.Figure: lambda _: None}, ttl=3600)
def barplot_suma(filtered_df):
    if len(selected_columns)==1:
        st.write('Proszę zaznaczyć co najmniej jedno medium, aby wyświetlić wykres z sumą.')
        st.stop()
    top10 = filtered_df[selected_columns[:-1]]
    top10 = top10.applymap(lambda x: int(x.replace(' ', '').replace('-', '0')))
    top10['Suma_Selected'] = top10.sum(axis=1)
    top10= top10.sort_values('Suma_Selected', ascending=False).head(10)

    bottom = np.zeros(len(top10))
    fig, ax = plt.subplots(figsize=(8, 6))
    for column in top10.columns[:-1]:
        bars = plt.barh(top10.index, top10[column], left=bottom, label=column, color=my_colors[column], height=0.5)
        bottom += top10[column]

    plt.title('Top 10 Suma', loc='left', fontdict={'fontsize': 14, 'fontweight': 'bold', 'fontname': 'Lato'})
    plt.gca().invert_yaxis()
    if selected_typ!=['Dzienniki regionalne']:
        plt.legend(loc=(0.8, 0.15))
    else:
        plt.legend(loc=(1.05, 0.15))

    plt.gca().spines[:].set_visible(False)
    plt.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False)
    plt.tick_params(axis='y', which='both', length=0, labelleft=False)

    margin = max(top10['Suma_Selected'])*0.02
    for bar, value in zip(bars, top10['Suma_Selected']):
        plt.text(bar.get_x()+bar.get_width()+margin,
                bar.get_y() + bar.get_height() / 2,
                format_number_with_spaces(value),
                ha='left',
                va='center'
                )
        
    for index, pismo in enumerate(list(top10.index)):
                plt.text(0, index-0.48, pismo, ha='left', va='center', fontdict={'fontsize': 10.8, 'fontname': 'Lato'})
    return fig


@st.cache_data(hash_funcs={matplotlib.figure.Figure: lambda _: None}, ttl=3600)
def create_donut(donutdf):
    sumdict = {}
    media = ['Facebook', 'X', 'YouTube', 'Instagram', 'LinkedIn', 'TikTok', 'Pinterest']
    for column in media+['Suma']:
        sumdict[column] = donutdf[column].sum()
    
    followers = [sumdict[medium] for medium in media]
    # rozwiązanie dla nachodzących na siebie podpisów (przy zmiane danych będzie wymagało korekty)
    label = [medium + ': ' + str(round(100*sumdict[medium]/sumdict['Suma'], 1)).replace('.',',')+'%' for medium in media[:4]] + ['']*3
    for i, medium in enumerate(media[4:]):
        plt.text(0.5*i-0.55, 1.05+(i%2)/11, medium + ': ' + str(round(100*sumdict[medium]/sumdict['Suma'], 1)).replace('.',',')+'%',
                horizontalalignment='center', verticalalignment='center', color='#31333f',
                fontdict={'fontsize': 8.8, 'fontname': 'Lato'})

    plt.pie(followers, colors=[my_colors[medium] for medium in media],
             labels=label, labeldistance=1.07,
             startangle=90, counterclock=False,
             textprops={'fontsize': 8.8, 'fontname': 'Lato', 'color': '#31333f'})#, explode=tuple([0.1] *len(Media)))

    centre_circle = plt.Circle((0, 0), 0.70, fc='white')
    fig = plt.gcf()
    fig.gca().add_artist(centre_circle)

    plt.text(0, .1, 'Suma obserwatorów',
            horizontalalignment='center', verticalalignment='center', color='#31333f',
            fontdict={'fontsize': 13, 'fontname': 'Lato', 'fontweight': 'bold'})
    plt.text(0, -.1, str(round(sumdict['Suma']/pow(10, 6), 1)).replace('.', ',') +' mln',
            horizontalalignment='center', verticalalignment='center', color='#31333f',
            fontdict={'fontsize': 27, 'fontname': 'Lato', 'fontweight': 'bold'})
    plt.axis('equal')
    return(fig)
######## Koniec wykresów ########


######## Wykres kołowy ########
st.markdown('---')
st.markdown("<h2 style='text-align: center; font-size: 1.27em;'>Udział poszczególnych platform</h2>", unsafe_allow_html=True)
st.pyplot(create_donut(donutdf))
###############################


st.markdown('---')
st.markdown("<h2 style='text-align: center; font-size: 1.27em;'>Liczba obserwatorów w social mediach</h2>", unsafe_allow_html=True)

typ = list(df['Typ'].unique())
selected_typ = st.multiselect('Wybierz grupę pism:', options=typ, default=typ)
if selected_typ:
    filtered_df = df[df['Typ'].isin(selected_typ)]
else:
     filtered_df = df

media=list(df.columns.drop(['Typ']))
st.markdown('<p style="font-size: 14px;">Wybierz media społecznościowe:</p>', unsafe_allow_html=True)
col1, col2 = st.columns(2)
with col1:
    facebook = st.checkbox('Facebook', value=True)
    instagram = st.checkbox('Instagram', value=True)
    linkedin = st.checkbox('LinkedIn', value=True)
    pinterest = st.checkbox('Pinterest', value=True)
    
with col2:
    youtube = st.checkbox('YouTube', value=True)
    x = st.checkbox('X', value=True)
    tiktok = st.checkbox('TikTok', value=True)
    suma = st.checkbox('Suma', value=True)

medialist = [
    (facebook, 'Facebook'),
    (youtube, 'YouTube'),
    (instagram, 'Instagram'),
    (x, 'X'),
    (linkedin, 'LinkedIn'),
    (tiktok, 'TikTok'),
    (pinterest, 'Pinterest'),
    (suma, 'Suma')
]
selected_columns = [x[1] for x in medialist if x[0]]
# Przypadek, gdy nic nie jest zaznaczone
if len(selected_columns)==0:
    selected_columns = [x[1] for x in medialist]

# Aktualizacja sumy w zależności do wybranych pism
if 'Suma' in selected_columns:
    if len(selected_columns)>1:
        filtered_df['Suma'] = filtered_df[selected_columns[:-1]].sum(axis=1)
        filtered_df = filtered_df.sort_values('Suma', ascending=False)
    else:
        filtered_df['Suma'] = df['Suma'] # suma dla wszystkich pism

filtered_df = filtered_df.replace(0, '-')
filtered_df = filtered_df[selected_columns].applymap(lambda x: format_number_with_spaces(x) if x != '-' else x)


# Sortowanie według 1 medium, jeśli tylko 1 medium wybrane
if len(selected_columns)==1:
    filtered_df[selected_columns[0]] = filtered_df[selected_columns[0]].str.replace(' ', '').replace('-', 0).astype(int)
    filtered_df.sort_values(by=selected_columns[0], ascending=False, inplace=True)
    filtered_df[selected_columns[0]] = filtered_df[selected_columns[0]].replace(0, '-').astype(str)
    filtered_df = filtered_df.applymap(lambda x: format_number_with_spaces(x) if x != '-' else x)


######## Wyświetlanie danych ########
output_type = st.radio('Wybierz tryb wyświetlania danych:', ['Tabela', 'Wykresy'], horizontal=True)
if output_type == 'Tabela':
    searchbar = st.text_input("Wyszukaj markę prasową:",  "", key="placeholder")
    filtered_df = filtered_df[filtered_df.index.str.contains(searchbar, case=False, na=False)]

    filtered_df.index.name = None
    filtered_df = filtered_df.rename(index=lambda x: add_hyperlink(x))

    filtered_df = filtered_df.style.set_table_styles([
        {'selector': 'table', 'props': [('text-align', 'center')]},
        {'selector': 'th', 'props': [('text-align', 'center')]},
        {'selector': 'td', 'props': [('text-align', 'center')]},
        {'selector': 'th.col0, td.col0', 'props': [('text-align', 'center')]}  # Update the selector for the first column
    ])

    filtered_df_html = filtered_df.to_html()
    filtered_df_html = filtered_df_html.replace('<table', "<table class='sticky-header'")

    st.markdown(table_css_style, unsafe_allow_html=True)
    st.markdown(f"<div class='sticky-table'>{filtered_df_html}</div>", unsafe_allow_html=True)

else:
    if selected_typ==['Dodatki']:
        st.write('Proszę wybrać więcej grup pism, kategoria jest zbyt wąska, by wyświetlać wykresy.')
        st.stop()

    if 'Suma' in selected_columns:
        st.pyplot(barplot_suma(filtered_df))
    
    for column in selected_columns:
        if column=='Suma':
            continue

        aux = filtered_df[filtered_df[column]!='-'][column]
        # Zamiana wartości na całkowite
        aux = aux.str.replace(' ', '').replace('-', 0).astype(int).sort_values(ascending=False).head(10)

        if len(aux)==0:
            st.write(f'Brak danych dla kategorii {column}')
            continue
        # przypadki gdzie jest mniej niż 10 pism w danej kategorii
        while len(aux)<10:
            aux = pd.concat([aux, pd.Series({' '*len(aux): 0})], axis=0)
        
        st.pyplot(barplot(df=aux, column=column))

footer_message = "Źródło: Liczba obserwatorów w mediach społecznościowych, opracowanie własne PBC, dane na dzień 01.10.2024"
st.markdown(f"""<div style="font-size:12px">{footer_message}</div>""", unsafe_allow_html=True)
st.write('\n')


# Pobieranie przefiltrowanych danych 
if output_type == 'Tabela':
    spreadsheet = openpyxl.load_workbook('template.xlsx')
    sheet = spreadsheet.active
    sheet['A3'] = f'Data wykonania raportu: {datetime.now().strftime("%d.%m.%Y")}'
    sheet['A5'] = 'Pismo'
    sheet['A5'].fill = PatternFill(start_color='00b0f0', end_color='00b0f0', fill_type='solid')

    output_df = pd.read_html(filtered_df_html)[0]

    for col_index, column in enumerate(list(output_df.columns)[1:], start=2):
        cell = sheet.cell(row=5, column=col_index, value=column)
        cell.fill = PatternFill(start_color='00b0f0', end_color='00b0f0', fill_type='solid')

    for row_index, row_data in enumerate(output_df.values, start=6):
        for col_index, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(horizontal='center')
            if row_index % 2 == 0:
                cell.fill = PatternFill(start_color='d9e1f2', end_color='d9e1f2', fill_type='solid')
    try:
        sheet.cell(row=row_index+1, column=1, value=footer_message)
    except NameError:
        pass

    output_file = 'Raport_Social_Media.xlsx'
    spreadsheet.save(output_file)

    st.download_button(label="Pobierz powyższą tabelę",
                        data=open(output_file, 'rb').read(),
                        file_name=output_file,
                            mime="application/vnd.ms-excel"
                            )


gc.collect()
